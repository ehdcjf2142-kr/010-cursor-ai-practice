"""
Read gold_prices_latest100.xlsx from 12_gold-price/, compute descriptive statistics,
and write a new Excel workbook under this folder.
"""

from __future__ import annotations

import statistics
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT.parent / "12_gold-price" / "gold_prices_latest100.xlsx"
OUT_XLSX = ROOT / "gold_price_stats_report.xlsx"


def parse_won(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def fmt_won(n: float | int) -> str:
    if isinstance(n, float):
        if n != n:  # NaN guard (should not happen)
            return ""
        rounded = round(n, 2)
        if rounded == int(rounded):
            return f"{int(rounded):,}"
        return f"{rounded:,.2f}"
    return f"{int(n):,}"


def load_numeric_columns(ws) -> tuple[list[str], list[list[int]]]:
    headers: list[str] = []
    for col in range(2, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        headers.append(str(h) if h else f"열{col}")

    series: list[list[int]] = [[] for _ in headers]
    for row in range(2, ws.max_row + 1):
        for i, col in enumerate(range(2, ws.max_column + 1)):
            v = parse_won(ws.cell(row=row, column=col).value)
            if v is not None:
                series[i].append(v)
    return headers, series


def load_dates(ws) -> list[str]:
    out: list[str] = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v is not None:
            out.append(str(v).strip())
    return out


def main() -> None:
    if not SOURCE_XLSX.is_file():
        raise SystemExit(f"원본 파일이 없습니다: {SOURCE_XLSX}")

    wb_in = load_workbook(SOURCE_XLSX, data_only=True)
    ws_in = wb_in.active

    headers, columns = load_numeric_columns(ws_in)
    dates = load_dates(ws_in)
    n = len(columns[0]) if columns else 0

    stat_rows = [
        ("표본 수 (건)", lambda col: len(col)),
        ("평균 (원)", lambda col: statistics.mean(col) if col else float("nan")),
        ("중앙값 (원)", lambda col: statistics.median(col) if col else float("nan")),
        ("최소 (원)", lambda col: min(col) if col else float("nan")),
        ("최대 (원)", lambda col: max(col) if col else float("nan")),
        (
            "표준편차 (원, 표본)",
            lambda col: statistics.stdev(col) if len(col) > 1 else 0.0,
        ),
    ]

    wb = Workbook()

    # --- Sheet: 메타 + 통계표 ---
    ws_meta = wb.active
    ws_meta.title = "요약·통계"

    ws_meta["A1"] = "분석 대상"
    ws_meta["B1"] = str(SOURCE_XLSX)
    ws_meta["A2"] = "분석 시각"
    ws_meta["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta["A3"] = "데이터 행 수"
    ws_meta["B3"] = n
    if dates:
        ws_meta["A4"] = "고시날짜 범위"
        ws_meta["B4"] = f"{min(dates)} ~ {max(dates)}"

    for addr in ("A1", "A2", "A3", "A4"):
        if ws_meta[addr].value:
            ws_meta[addr].font = Font(bold=True)

    start_row = 6
    ws_meta.cell(row=start_row, column=1, value="통계 항목")
    ws_meta.cell(row=start_row, column=1).font = Font(bold=True)
    for j, h in enumerate(headers, start=2):
        c = ws_meta.cell(row=start_row, column=j, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, (label, fn) in enumerate(stat_rows, start=start_row + 1):
        ws_meta.cell(row=i, column=1, value=label)
        ws_meta.cell(row=i, column=1).alignment = Alignment(horizontal="left")
        for j, col in enumerate(columns, start=2):
            val = fn(col)
            if label.startswith("표본 수"):
                ws_meta.cell(row=i, column=j, value=val)
            else:
                display = fmt_won(val) if isinstance(val, (int, float)) else val
                ws_meta.cell(row=i, column=j, value=display)
                ws_meta.cell(row=i, column=j).alignment = Alignment(horizontal="right")

    ws_meta.column_dimensions["A"].width = 28
    for j in range(2, len(headers) + 2):
        ws_meta.column_dimensions[get_column_letter(j)].width = 26

    # --- Sheet: 원본 복사 (확인용) ---
    ws_copy = wb.create_sheet("원본데이터_복사")
    for r in range(1, ws_in.max_row + 1):
        for c in range(1, ws_in.max_column + 1):
            ws_copy.cell(row=r, column=c, value=ws_in.cell(row=r, column=c).value)
    if ws_copy.max_row >= 1:
        for c in ws_copy[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    wb.save(OUT_XLSX)
    print(f"통계 엑셀 저장: {OUT_XLSX}")


if __name__ == "__main__":
    main()
