"""
Fetch domestic gold prices from Korea Gold Exchange (same source as the Tabulator on /price/gold)
and save the latest 100 rows to Excel.
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

API_URL = "https://www.koreagoldx.co.kr/api/price/chart/list"
SOURCE_PAGE = "https://www.koreagoldx.co.kr/price/gold"
ROW_LIMIT = 100


def fetch_list(session: requests.Session) -> list[dict]:
    end = datetime.now()
    start = end - timedelta(days=400)
    payload = {
        "srchDt": "SEARCH",
        "type": "Au",
        "dataDateStart": start.strftime("%Y.%m.%d"),
        "dataDateEnd": end.strftime("%Y.%m.%d"),
    }
    r = session.post(
        API_URL,
        data=json.dumps(payload, ensure_ascii=False),
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "Accept": "application/json",
            "Referer": SOURCE_PAGE,
            "Origin": "https://www.koreagoldx.co.kr",
        },
        timeout=120,
    )
    r.raise_for_status()
    data = r.json()
    return data.get("list") or []


def format_display_date(raw: str) -> str:
    """API uses 'YYYY-MM-DD HH:mm:ss'; site table shows 'YYYY.MM.DD'."""
    raw = (raw or "").strip()
    if len(raw) >= 10:
        y, m, d = raw[:10].split("-")
        return f"{y}.{m}.{d}"
    return raw


def money_won(n: int | None) -> str | int:
    if n is None:
        return ""
    return f"{n:,}"


def main() -> None:
    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / "gold_prices_latest100.xlsx"

    session = requests.Session()
    session.headers["User-Agent"] = (
        "Mozilla/5.0 (compatible; GoldPriceExport/1.0; +local script)"
    )

    rows = fetch_list(session)
    rows = rows[:ROW_LIMIT]

    wb = Workbook()
    ws = wb.active
    ws.title = "금시세"

    headers = [
        "고시날짜",
        "내가 살 때(3.75g) 순금 (원)",
        "내가 팔 때(3.75g) 순금 (원)",
        "내가 팔 때(3.75g) 18K (원)",
        "내가 팔 때(3.75g) 14K (원)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in rows:
        ws.append(
            [
                format_display_date(item.get("date", "")),
                money_won(item.get("s_pure")),
                money_won(item.get("p_pure")),
                money_won(item.get("p_18k")),
                money_won(item.get("p_14k")),
            ]
        )

    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

    wb.save(out_path)
    print(f"Saved {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    main()
